"""Generates Excel risk dashboard and matplotlib chart from analysis results."""

import json
import pathlib
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

REPORTS = pathlib.Path("reports")

RED = "FFCCCC"
YELLOW = "FFF3CC"
GREEN = "CCFFCC"
HEADER = "2B5797"


def _risk_color(score: float) -> str:
    if score >= 0.7:
        return RED
    if score >= 0.4:
        return YELLOW
    return GREEN


def write_excel(results: dict, df: pd.DataFrame, out_path: pathlib.Path) -> None:
    wb = Workbook()

    # --- Sheet 1: Executive Summary ---
    ws = wb.active
    ws.title = "Executive Summary"

    headers = ["Site", "Total Incidents", "Risk Score", "Trend",
               "INTRUSION", "ALARM", "ACCESS_FAIL", "TAILGATING"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center")

    for row, (site, data) in enumerate(sorted(results.items()), start=2):
        score = data["risk_score"]
        row_data = [
            site,
            data["total"],
            score,
            data["trend"],
            data["by_type"].get("INTRUSION", 0),
            data["by_type"].get("ALARM", 0),
            data["by_type"].get("ACCESS_FAIL", 0),
            data["by_type"].get("TAILGATING", 0),
        ]
        fill = PatternFill("solid", fgColor=_risk_color(score))
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # --- Sheet 2: Incident Log (first 500 rows) ---
    ws2 = wb.create_sheet("Incident Log")
    log_cols = list(df.columns)
    for col, h in enumerate(log_cols, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER)
    for row, record in enumerate(df.head(500).itertuples(index=False), start=2):
        for col, val in enumerate(record, 1):
            ws2.cell(row=row, column=col, value=val)
    for col in range(1, len(log_cols) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    wb.save(out_path)


def write_chart(results: dict, out_path: pathlib.Path) -> None:
    sites = sorted(results.keys())
    scores = [results[s]["risk_score"] for s in sites]
    colors = [
        "#CC4444" if s >= 0.7 else "#CCAA00" if s >= 0.4 else "#44AA44"
        for s in scores
    ]

    fig, ax = plt.subplots(figsize=(8, 4))
    bars = ax.bar(sites, scores, color=colors)
    ax.set_ylim(0, 1.0)
    ax.axhline(0.7, color="#CC4444", linestyle="--", linewidth=0.8, label="High risk threshold")
    ax.axhline(0.4, color="#CCAA00", linestyle="--", linewidth=0.8, label="Medium risk threshold")
    ax.set_xlabel("Site")
    ax.set_ylabel("Risk Score (0–1)")
    ax.set_title("Per-Site Security Risk Scores")
    ax.legend(fontsize=8)
    for bar, score in zip(bars, scores):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02,
                f"{score:.2f}", ha="center", fontsize=9)
    plt.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close()


def write_benchmark(data: dict, out_path: pathlib.Path) -> None:
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)
